from flask import Flask, render_template, request, jsonify
import os
import csv
import numpy as np

# Constants
BASE_TO_INT = {'A': 0, 'C': 1, 'G': 2, 'T': 3, 'N': 4}
MAX_LEN = 1170

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Load model weights
npz_path = os.path.join(os.path.dirname(__file__), 'enhancer_model_weights.npz')
try:
    weights = np.load(npz_path)
    print("Successfully loaded NumPy model weights.")
except Exception as e:
    print(f"Warning: Could not load weights from {npz_path}: {e}")
    weights = None

def sigmoid(x):
    # Clip to avoid overflow/underflow warnings
    x = np.clip(x, -500, 500)
    return 1 / (1 + np.exp(-x))

def numpy_predict(x):
    # x is a NumPy array of shape (batch_size, MAX_LEN)
    if weights is None:
        print("NumPy weights not loaded. Returning default prediction.")
        return np.full(x.shape[0], 0.5)
        
    try:
        embed_w = weights['embedding.weight']
        conv_w = weights['conv1.weight']
        conv_b = weights['conv1.bias']
        
        lstm_w_ih = weights['lstm.weight_ih_l0']
        lstm_w_hh = weights['lstm.weight_hh_l0']
        lstm_b_ih = weights['lstm.bias_ih_l0']
        lstm_b_hh = weights['lstm.bias_hh_l0']
        
        lstm_w_ih_rev = weights['lstm.weight_ih_l0_reverse']
        lstm_w_hh_rev = weights['lstm.weight_hh_l0_reverse']
        lstm_b_ih_rev = weights['lstm.bias_ih_l0_reverse']
        lstm_b_hh_rev = weights['lstm.bias_hh_l0_reverse']
        
        fc1_w = weights['fc1.weight']
        fc1_b = weights['fc1.bias']
        fc2_w = weights['fc2.weight']
        fc2_b = weights['fc2.bias']
        
        batch_size, seq_len = x.shape
        
        # 1. Embedding -> (batch, seq_len, 16)
        embed = embed_w[x]
        
        # 2. Conv1D with padding=2
        # Pad sequence dimension (axis 1) by 2 on left and 2 on right
        embed_padded = np.pad(embed, ((0, 0), (2, 2), (0, 0)), mode='constant', constant_values=0)
        L_out = seq_len
        conv_out = np.zeros((batch_size, L_out, 32))
        conv_w_flat = conv_w.reshape(32, -1)
        
        for t in range(L_out):
            window = embed_padded[:, t:t+5, :].transpose(0, 2, 1).reshape(batch_size, -1)
            conv_out[:, t, :] = np.dot(window, conv_w_flat.T) + conv_b
            
        conv_out = np.maximum(conv_out, 0)  # ReLU
        
        # 3. MaxPool1D (stride=8, kernel=8)
        L_pool = L_out // 8
        conv_out_truncated = conv_out[:, :L_pool * 8, :]
        reshaped = conv_out_truncated.reshape(batch_size, L_pool, 8, 32)
        pooled = np.max(reshaped, axis=2)  # shape: (batch, L_pool, 32)
        
        # 4. BiLSTM
        hidden_size = 64
        
        # Forward LSTM
        h_f = np.zeros((batch_size, hidden_size))
        c_f = np.zeros((batch_size, hidden_size))
        bias_f = lstm_b_ih + lstm_b_hh
        
        for t in range(L_pool):
            x_t = pooled[:, t, :]
            gates = np.dot(x_t, lstm_w_ih.T) + np.dot(h_f, lstm_w_hh.T) + bias_f
            i = sigmoid(gates[:, :hidden_size])
            f = sigmoid(gates[:, hidden_size:2*hidden_size])
            g = np.tanh(gates[:, 2*hidden_size:3*hidden_size])
            o = sigmoid(gates[:, 3*hidden_size:])
            c_f = f * c_f + i * g
            h_f = o * np.tanh(c_f)
            
        # Backward LSTM
        h_b = np.zeros((batch_size, hidden_size))
        c_b = np.zeros((batch_size, hidden_size))
        bias_b = lstm_b_ih_rev + lstm_b_hh_rev
        
        for t in reversed(range(L_pool)):
            x_t = pooled[:, t, :]
            gates = np.dot(x_t, lstm_w_ih_rev.T) + np.dot(h_b, lstm_w_hh_rev.T) + bias_b
            i = sigmoid(gates[:, :hidden_size])
            f = sigmoid(gates[:, hidden_size:2*hidden_size])
            g = np.tanh(gates[:, 2*hidden_size:3*hidden_size])
            o = sigmoid(gates[:, 3*hidden_size:])
            c_b = f * c_b + i * g
            h_b = o * np.tanh(c_b)
            
        # Concatenate forward and backward final states -> (batch, 128)
        hidden = np.concatenate((h_f, h_b), axis=1)
        
        # 5. FC1
        fc1_out = np.dot(hidden, fc1_w.T) + fc1_b
        fc1_out = np.maximum(fc1_out, 0)  # ReLU
        
        # 6. FC2
        logits = np.dot(fc1_out, fc2_w.T) + fc2_b
        probs = sigmoid(logits)
        
        return probs.squeeze(axis=-1)
    except Exception as e:
        print("Prediction error in NumPy forward pass:", e)
        return np.full(x.shape[0], 0.5)

def predict_sequence(seq):
    seq = seq.upper()
    seq_encoded = np.full(MAX_LEN, 4, dtype=np.int64)
    for i, base in enumerate(seq):
        if i >= MAX_LEN: break
        seq_encoded[i] = BASE_TO_INT.get(base, 4)
    
    # Run prediction using NumPy
    batch = np.expand_dims(seq_encoded, axis=0)
    probs = numpy_predict(batch)
    return float(probs[0])

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/predict_text', methods=['POST'])
def predict_text():
    data = request.get_json()
    sequence = data.get('sequence', '').strip()
    if not sequence:
        return jsonify({'error': 'No sequence provided'}), 400
    
    prob = predict_sequence(sequence)
    is_enhancer = prob > 0.5
    
    return jsonify({
        'prediction': 'Regulatory Element (Enhancer)' if is_enhancer else 'Non-Regulatory',
        'confidence': f"{(prob if is_enhancer else 1-prob) * 100:.2f}%",
        'raw_probability': prob
    })

@app.route('/predict_csv', methods=['POST'])
def predict_csv():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'})
        
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)
    
    try:
        rows = []
        if filepath.endswith('.csv'):
            # Read CSV using built-in csv module with header auto-scanning
            with open(filepath, mode='r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                all_csv_rows = list(reader)
                
            # Scan the first few rows to locate the header row
            header_row_idx = None
            for i, row in enumerate(all_csv_rows[:10]):
                row_str_vals = [str(val).lower() for val in row if val is not None]
                if any('sequence' in val or 'seq' in val or 'class' in val or 'label' in val for val in row_str_vals):
                    header_row_idx = i
                    break
            
            if header_row_idx is None:
                header_row_idx = 0
                
            headers = all_csv_rows[header_row_idx]
            
            # Parse subsequent rows
            for row_cells in all_csv_rows[header_row_idx + 1:]:
                row_dict = {}
                for header, val in zip(headers, row_cells):
                    if header:
                        row_dict[header] = val
                rows.append(row_dict)
                if len(rows) >= 150:
                    break
                    
        elif filepath.endswith('.xlsx'):
            # Read Excel using openpyxl directly with header auto-scanning
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet = wb.active
            
            # Scan the first few rows to locate the header row
            rows_data = list(sheet.iter_rows(max_row=10, values_only=True))
            header_row_idx = None
            for i, row in enumerate(rows_data):
                row_str_vals = [str(val).lower() for val in row if val is not None]
                if any('sequence' in val or 'seq' in val or 'class' in val or 'label' in val for val in row_str_vals):
                    header_row_idx = i
                    break
            
            if header_row_idx is None:
                header_row_idx = 0
            
            # Extract headers from the identified header row
            headers = []
            for cell in rows_data[header_row_idx]:
                headers.append(str(cell) if cell is not None else "")
            
            # Read subsequent rows starting from the row after headers
            row_count = 0
            for row_cells in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
                row_dict = {}
                for header, val in zip(headers, row_cells):
                    if header:
                        row_dict[header] = val if val is not None else ""
                rows.append(row_dict)
                row_count += 1
                if row_count >= 150:
                    break
            wb.close()
        else:
            return jsonify({'error': 'Invalid file type. Please upload a CSV or XLSX file.'}), 400
            
        if not rows:
            return jsonify({'error': 'No data found in file.'}), 400
            
        columns = list(rows[0].keys())
        
        # Find combination of columns
        seq_col = None
        for col in columns:
            if 'seq' in col.lower() or 'sequence' in col.lower():
                seq_col = col
                break
        
        if not seq_col:
            # Fallback to the second column
            if len(columns) >= 2:
                seq_col = columns[1]
            else:
                return jsonify({'error': 'Could not identify a Sequence column. Please ensure there is a column named "sequence".'}), 400
                
        class_col = None
        for col in columns:
            if 'class' in col.lower() or 'label' in col.lower():
                class_col = col
                break
                
        # Collect up to 100 valid sequences
        sequences_to_predict = []
        rows_to_process = []
        
        for row in rows:
            if len(rows_to_process) >= 100:
                break
                
            seq = str(row.get(seq_col, '')).strip()
            # If the file includes headers inside the data occasionally
            if seq.lower() == 'sequence': continue 
            
            if len(seq) < 10: continue
            
            rows_to_process.append(row)
            sequences_to_predict.append(seq)
            
        # Perform batched NumPy prediction
        probs = []
        if sequences_to_predict:
            encoded_list = []
            for seq in sequences_to_predict:
                seq = seq.upper()
                seq_encoded = np.full(MAX_LEN, 4, dtype=np.int64)
                for i, base in enumerate(seq):
                    if i >= MAX_LEN: break
                    seq_encoded[i] = BASE_TO_INT.get(base, 4)
                encoded_list.append(seq_encoded)
                
            try:
                # Batch of size N: shape (N, MAX_LEN)
                batch = np.array(encoded_list, dtype=np.int64)
                probs = numpy_predict(batch).tolist()
            except Exception as e:
                print("Batch prediction error in NumPy runtime:", e)
                probs = [0.5] * len(sequences_to_predict)
        else:
            probs = [0.5] * len(sequences_to_predict)
                    
        # Construct results
        results = []
        correct = 0
        total = 0
        
        for i, row in enumerate(rows_to_process):
            seq = sequences_to_predict[i]
            prob = probs[i]
            pred_label = 1 if prob > 0.5 else 0
            
            res_dict = {
                'sequence_preview': seq[:30] + '...',
                'full_sequence': seq,
                'prediction': 'Enhancer' if pred_label == 1 else 'Non-Enhancer',
                'confidence': f"{(prob if pred_label == 1 else 1-prob) * 100:.2f}%"
            }
            
            val = row.get(class_col)
            if class_col and val is not None and str(val).strip() != "":
                try: # try parsing label
                    actual_label = int(float(val))
                    res_dict['actual'] = 'Enhancer' if actual_label == 1 else 'Non-Enhancer'
                    if pred_label == actual_label:
                        correct += 1
                    total += 1
                except:
                    pass
            results.append(res_dict)
            
        accuracy = None
        if total > 0:
            accuracy = f"{(correct / total) * 100:.2f}%"
            
        return jsonify({
            'results': results,
            'accuracy': accuracy,
            'total_processed': len(results)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
